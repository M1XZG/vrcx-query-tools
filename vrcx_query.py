#!/usr/bin/env python3
"""
VRCX Database Query Script
Demonstrates how to access the live VRCX database and analyze VRChat instance data.

This script provides examples for:
1. Connecting to the VRCX SQLite database
2. Querying location and join/leave events
3. Creating an hour-by-hour report of people in instances
4. Exporting data to CSV/Excel

Requirements:
    pip install sqlite3 pandas openpyxl
    (sqlite3 is built-in with Python)

Usage:
    python vrcx_query.py
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
import json
import argparse

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    # If python-dotenv is not installed, manually load .env file
    env_path = Path(__file__).parent / '.env'
    if env_path.exists():
        with open(env_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()

# ==============================================================================
# Configuration
# ==============================================================================

# Try to find the VRCX database automatically
def find_vrcx_database():
    """Locate the VRCX.sqlite3 database file."""
    # Windows
    if sys.platform == 'win32':
        appdata = os.getenv('APPDATA')
        if appdata:
            db_path = Path(appdata) / 'VRCX' / 'VRCX.sqlite3'
            if db_path.exists():
                return str(db_path)
    
    # Linux/Mac
    home = Path.home()
    possible_paths = [
        home / '.config' / 'VRCX' / 'VRCX.sqlite3',  # Linux
        home / 'Library' / 'Application Support' / 'VRCX' / 'VRCX.sqlite3',  # macOS
    ]
    
    for path in possible_paths:
        if path.exists():
            return str(path)
    
    return None


DATABASE_PATH = os.getenv('VRCX_DATABASE_PATH') or find_vrcx_database()

if not DATABASE_PATH:
    print("ERROR: Could not find VRCX.sqlite3 database")
    print("Please set VRCX_DATABASE_PATH environment variable or ensure VRCX is installed")
    sys.exit(1)

print(f"Using database: {DATABASE_PATH}")

# ==============================================================================
# Database Connection
# ==============================================================================

class VRCXDatabase:
    """Interface to the VRCX SQLite database."""
    
    def __init__(self, db_path):
        self.db_path = db_path
        self.connection = None
    
    def connect(self):
        """Establish database connection."""
        try:
            self.connection = sqlite3.connect(self.db_path)
            self.connection.row_factory = sqlite3.Row  # Return rows as dictionaries
            print(f"[OK] Connected to database")
        except sqlite3.Error as e:
            print(f"✗ Database connection failed: {e}")
            sys.exit(1)
    
    def close(self):
        """Close database connection."""
        if self.connection:
            self.connection.close()
            print(f"[OK] Database connection closed")
    
    def execute(self, query, params=None):
        """Execute a query and return results."""
        if not self.connection:
            self.connect()
        
        try:
            cursor = self.connection.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchall()
        except sqlite3.Error as e:
            print(f"✗ Query failed: {e}")
            print(f"Query: {query}")
            if params:
                print(f"Params: {params}")
            return []
    
    def get_table_names(self):
        """Get list of all tables in database."""
        query = "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        rows = self.execute(query)
        return [row['name'] for row in rows]


# ==============================================================================
# Query Functions
# ==============================================================================

class VRCXQuery:
    """Query helper for common VRCX analysis tasks."""
    
    def __init__(self, db):
        self.db = db
    
    def get_location_history(self, date_str=None):
        """
        Get your location/instance history for a specific date.
        
        Args:
            date_str: Date in format 'YYYY-MM-DD'. If None, uses today.
        
        Returns:
            List of location entries with timestamps and duration.
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        query = """
        SELECT 
            id,
            created_at,
            location,
            world_id,
            world_name,
            time as duration_seconds,
            group_name
        FROM gamelog_location
        WHERE DATE(created_at) = ?
        ORDER BY created_at ASC
        """
        
        results = self.db.execute(query, (date_str,))
        return results
    
    def get_join_leave_events(self, location=None, date_str=None):
        """
        Get join/leave events for a specific location and date.
        
        Args:
            location: Instance ID (e.g., 'wrld_12345:12345~region(us)') or None for all
            date_str: Date in format 'YYYY-MM-DD'. If None, uses today.
        
        Returns:
            List of join/leave events with user info.
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        if location:
            query = """
            SELECT 
                id,
                created_at,
                type,
                display_name,
                user_id,
                location,
                time
            FROM gamelog_join_leave
            WHERE location = ? AND DATE(created_at) = ?
            ORDER BY created_at ASC
            """
            results = self.db.execute(query, (location, date_str))
        else:
            query = """
            SELECT 
                id,
                created_at,
                type,
                display_name,
                user_id,
                location,
                time
            FROM gamelog_join_leave
            WHERE DATE(created_at) = ?
            ORDER BY created_at ASC
            """
            results = self.db.execute(query, (date_str,))
        
        return results
    
    def get_hour_by_hour_summary(self, date_str=None):
        """
        Get hour-by-hour summary of instances and people for a specific date.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        
        Returns data like:
        Hour | Unique People
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(j.display_name) as unique_people
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (date_str, current_hour))
        return results
    
    def get_hour_by_hour_average(self, start_date_str=None, end_date_str=None):
        """
        Get average hour-by-hour attendance across a date range.
        Counts total join/leave events per hour.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        
        Returns average data like:
        Hour | Avg People
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        ),
        hourly_data AS (
            SELECT 
                DATE(created_at) as date,
                CAST(strftime('%H', created_at) AS INTEGER) as hour,
                COUNT(*) as total_people
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at), hour
        )
        SELECT 
            h.hour,
            CAST(ROUND(AVG(COALESCE(hd.total_people, 0))) AS INTEGER) as avg_unique_people
        FROM hours h
        LEFT JOIN hourly_data hd ON h.hour = hd.hour
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_daily_hourly_summary(self, start_date_str=None, end_date_str=None):
        """
        Get hour-by-hour summary for each day in a date range.
        Counts total join/leave events.
        
        Returns data like:
        Date | Hour | People
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            DATE(created_at) as date,
            CAST(strftime('%H', created_at) AS INTEGER) as hour,
            COUNT(*) as unique_people
        FROM gamelog_join_leave
        WHERE DATE(created_at) BETWEEN ? AND ?
        GROUP BY DATE(created_at), hour
        ORDER BY DATE(created_at) ASC, hour ASC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_day_of_week_average(self, start_date_str=None, end_date_str=None):
        """
        Get average attendance by day of week.
        Counts total join/leave events per day.
        
        Returns data like:
        Day of Week | Avg People
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            CAST(strftime('%w', date) AS INTEGER) as day_of_week,
            ROUND(AVG(total_people)) as avg_unique_people
        FROM (
            SELECT 
                DATE(created_at) as date,
                COUNT(*) as total_people
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at)
        )
        GROUP BY day_of_week
        ORDER BY day_of_week
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_weekly_day_of_week_breakdown(self, start_date_str=None, end_date_str=None):
        """
        Get attendance by day of week, grouped by week.
        Counts total join/leave events per day.
        
        Returns data like:
        Week Start | Week End | Day of Week | People
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            week_start,
            week_end,
            day_of_week,
            day_name,
            unique_people
        FROM (
            SELECT 
                DATE(created_at) as date,
                CAST(strftime('%w', created_at) AS INTEGER) as day_of_week,
                CASE CAST(strftime('%w', created_at) AS INTEGER)
                    WHEN 0 THEN 'Sunday'
                    WHEN 1 THEN 'Monday'
                    WHEN 2 THEN 'Tuesday'
                    WHEN 3 THEN 'Wednesday'
                    WHEN 4 THEN 'Thursday'
                    WHEN 5 THEN 'Friday'
                    WHEN 6 THEN 'Saturday'
                END as day_name,
                DATE(created_at, 'weekday 0', '-6 days') as week_start,
                DATE(created_at, 'weekday 0') as week_end,
                COUNT(*) as unique_people
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at)
        )
        GROUP BY week_start, week_end, day_of_week, day_name, date
        ORDER BY week_start, day_of_week
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_people_in_instances_by_hour(self, date_str=None):
        """
        Get detailed breakdown of who was in instances hour-by-hour.
        
        This version tracks individual people across the day.
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        query = """
        SELECT 
            CAST(strftime('%H', created_at) AS INTEGER) as hour,
            location,
            world_name,
            type,
            display_name,
            user_id,
            created_at
        FROM gamelog_join_leave
        WHERE DATE(created_at) = ?
        ORDER BY hour ASC, location ASC, created_at ASC
        """
        
        results = self.db.execute(query, (date_str,))
        return results
    
    def get_instance_statistics(self, date_str=None):
        """
        Get statistics about instances visited.
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        query = """
        SELECT 
            location,
            world_name,
            COUNT(*) as visits,
            SUM(time) as total_time_seconds,
            MIN(created_at) as first_visit,
            MAX(created_at) as last_visit
        FROM gamelog_location
        WHERE DATE(created_at) = ?
        GROUP BY location
        ORDER BY total_time_seconds DESC
        """
        
        results = self.db.execute(query, (date_str,))
        return results
    
    def get_unique_visitors_by_hour(self, date_str=None):
        """
        Get hour-by-hour count of unique visitors (each person counted once per hour).
        Each person is counted only once per hour, regardless of how many times they joined/left.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        
        Returns data like:
        Hour | Unique Visitors
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(DISTINCT j.display_name) as unique_visitors
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (date_str, current_hour))
        return results
    
    def get_unique_visitors_daily(self, start_date_str=None, end_date_str=None):
        """
        Get unique visitors per day across a date range.
        Each person is counted only once per day, regardless of how many times they joined/left.
        
        Returns data like:
        Date | Unique Visitors
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            DATE(created_at) as date,
            COUNT(DISTINCT display_name) as unique_visitors
        FROM gamelog_join_leave
        WHERE DATE(created_at) BETWEEN ? AND ?
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at) ASC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_unique_visitors_average(self, start_date_str=None, end_date_str=None):
        """
        Get average unique visitors per hour across a date range.
        Each person is counted only once per hour per day.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        
        Returns data like:
        Hour | Avg Unique Visitors
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        ),
        hourly_data AS (
            SELECT 
                DATE(created_at) as date,
                CAST(strftime('%H', created_at) AS INTEGER) as hour,
                COUNT(DISTINCT display_name) as unique_visitors
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at), hour
        )
        SELECT 
            h.hour,
            CAST(ROUND(AVG(COALESCE(hd.unique_visitors, 0))) AS INTEGER) as avg_unique_visitors
        FROM hours h
        LEFT JOIN hourly_data hd ON h.hour = hd.hour
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_unique_visitors_day_of_week(self, start_date_str=None, end_date_str=None):
        """
        Get average unique visitors by day of week across a date range.
        Each person is counted only once per day.
        
        Returns data like:
        Day of Week | Avg Unique Visitors
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            CAST(strftime('%w', date) AS INTEGER) as day_of_week,
            ROUND(AVG(unique_visitors)) as avg_unique_visitors
        FROM (
            SELECT 
                DATE(created_at) as date,
                COUNT(DISTINCT display_name) as unique_visitors
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at)
        )
        GROUP BY day_of_week
        ORDER BY day_of_week
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_unique_visitors_weekly(self, start_date_str=None, end_date_str=None):
        """
        Get unique visitors by day of week, grouped by week.
        Each person is counted only once per day.
        
        Returns data like:
        Week Start | Week End | Day of Week | Unique Visitors
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            week_start,
            week_end,
            day_of_week,
            day_name,
            unique_visitors
        FROM (
            SELECT 
                DATE(created_at) as date,
                CAST(strftime('%w', created_at) AS INTEGER) as day_of_week,
                CASE CAST(strftime('%w', created_at) AS INTEGER)
                    WHEN 0 THEN 'Sunday'
                    WHEN 1 THEN 'Monday'
                    WHEN 2 THEN 'Tuesday'
                    WHEN 3 THEN 'Wednesday'
                    WHEN 4 THEN 'Thursday'
                    WHEN 5 THEN 'Friday'
                    WHEN 6 THEN 'Saturday'
                END as day_name,
                DATE(created_at, 'weekday 0', '-6 days') as week_start,
                DATE(created_at, 'weekday 0') as week_end,
                COUNT(DISTINCT display_name) as unique_visitors
            FROM gamelog_join_leave
            WHERE DATE(created_at) BETWEEN ? AND ?
            GROUP BY DATE(created_at)
        )
        GROUP BY week_start, week_end, day_of_week, day_name, date
        ORDER BY week_start, day_of_week
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_unique_worlds(self, start_date_str=None, end_date_str=None):
        """
        Get list of unique worlds visited during a date range.
        
        Returns data like:
        World ID | World Name | Visit Count
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            world_id,
            world_name,
            COUNT(DISTINCT DATE(created_at)) as visit_count,
            COUNT(*) as total_events
        FROM gamelog_location
        WHERE DATE(created_at) BETWEEN ? AND ? AND world_id IS NOT NULL
        GROUP BY world_id, world_name
        ORDER BY visit_count DESC, world_name ASC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str))
        return results
    
    def get_unique_instances_for_world(self, world_id, start_date_str=None, end_date_str=None):
        """
        Get list of unique instances (with instance IDs) visited for a specific world during a date range.
        
        Args:
            world_id: World ID to filter by
            start_date_str: Start date
            end_date_str: End date
        
        Returns data like:
        Instance ID | Visit Count | Last Visited
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            location as instance_id,
            SUBSTR(location, INSTR(location, ':') + 1) as instance_number,
            COUNT(DISTINCT DATE(created_at)) as visit_count,
            MAX(created_at) as last_visited
        FROM gamelog_join_leave
        WHERE DATE(created_at) BETWEEN ? AND ? AND SUBSTR(location, 1, ?) = ?
        GROUP BY location
        ORDER BY visit_count DESC, last_visited DESC
        """
        
        world_id_len = len(world_id)
        results = self.db.execute(query, (start_date_str, end_date_str, world_id_len, world_id))
        return results
    
    def get_hour_by_hour_summary_for_instance(self, instance_id, date_str=None):
        """
        Get hour-by-hour summary for a specific instance on a given date.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        
        Args:
            instance_id: Full instance ID (location field value)
            date_str: Date in format 'YYYY-MM-DD'
        
        Returns:
            Hour-by-hour data for the specific instance
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(j.display_name) as unique_people
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? 
            AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
            AND j.location = ?
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (date_str, instance_id, current_hour))
        return results
    
    def get_unique_visitors_by_hour_for_instance(self, instance_id, date_str=None):
        """
        Get hour-by-hour count of unique visitors for a specific instance.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        
        Args:
            instance_id: Full instance ID (location field value)
            date_str: Date in format 'YYYY-MM-DD'
        
        Returns:
            Hour-by-hour unique visitor data for the specific instance
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(DISTINCT j.display_name) as unique_visitors
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? 
            AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
            AND j.location = ?
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        results = self.db.execute(query, (date_str, instance_id, current_hour))
        return results
    
    def get_users_for_instance(self, instance_id, start_date_str=None, end_date_str=None):
        """
        Get list of all unique users who visited a specific instance during a date range.
        
        Args:
            instance_id: Full instance ID (location field value)
            start_date_str: Start date
            end_date_str: End date
        
        Returns data like:
        User Name | Visit Count | Days Visited | First Visit | Last Visit
        """
        if start_date_str is None:
            start_date_str = datetime.now().strftime('%Y-%m-%d')
        if end_date_str is None:
            end_date_str = start_date_str
        
        query = """
        SELECT 
            display_name,
            COUNT(*) as visit_count,
            COUNT(DISTINCT DATE(created_at)) as days_visited,
            MIN(created_at) as first_visit,
            MAX(created_at) as last_visit
        FROM gamelog_join_leave
        WHERE DATE(created_at) BETWEEN ? AND ? AND location = ?
        GROUP BY display_name
        ORDER BY visit_count DESC, last_visit DESC
        """
        
        results = self.db.execute(query, (start_date_str, end_date_str, instance_id))
        return results
    
    def get_hour_by_hour_summary_for_world(self, world_id, date_str=None):
        """
        Get hour-by-hour summary for a specific world on a given date.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        Extracts world_id from location field (format: wrld_xxx:instance~...)
        
        Args:
            world_id: World ID to filter by (e.g., 'wrld_4432ea9b-729c-46e3-8eaf-846aa0a37fdd')
            date_str: Date in format 'YYYY-MM-DD'
        
        Returns:
            Hour-by-hour data for the specific world
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(j.display_name) as unique_people
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? 
            AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
            AND SUBSTR(j.location, 1, ?) = ?
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        # Extract world_id length for substring matching
        world_id_len = len(world_id)
        results = self.db.execute(query, (date_str, world_id_len, world_id, current_hour))
        return results
    
    def get_unique_visitors_by_hour_for_world(self, world_id, date_str=None):
        """
        Get hour-by-hour count of unique visitors for a specific world.
        Includes all 24 hours (0-23), with 0 for hours with no data.
        For today's date, only shows hours that have passed.
        Extracts world_id from location field (format: wrld_xxx:instance~...)
        
        Args:
            world_id: World ID to filter by (e.g., 'wrld_4432ea9b-729c-46e3-8eaf-846aa0a37fdd')
            date_str: Date in format 'YYYY-MM-DD'
        
        Returns:
            Hour-by-hour unique visitor data for the specific world
        """
        if date_str is None:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Check if this is today's date
        is_today = date_str == datetime.now().strftime('%Y-%m-%d')
        current_hour = datetime.now().hour if is_today else 23
        
        query = """
        WITH hours AS (
            SELECT 0 AS hour UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3
            UNION ALL SELECT 4 UNION ALL SELECT 5 UNION ALL SELECT 6 UNION ALL SELECT 7
            UNION ALL SELECT 8 UNION ALL SELECT 9 UNION ALL SELECT 10 UNION ALL SELECT 11
            UNION ALL SELECT 12 UNION ALL SELECT 13 UNION ALL SELECT 14 UNION ALL SELECT 15
            UNION ALL SELECT 16 UNION ALL SELECT 17 UNION ALL SELECT 18 UNION ALL SELECT 19
            UNION ALL SELECT 20 UNION ALL SELECT 21 UNION ALL SELECT 22 UNION ALL SELECT 23
        )
        SELECT 
            h.hour,
            COUNT(DISTINCT j.display_name) as unique_visitors
        FROM hours h
        LEFT JOIN gamelog_join_leave j ON DATE(j.created_at) = ? 
            AND CAST(strftime('%H', j.created_at) AS INTEGER) = h.hour
            AND SUBSTR(j.location, 1, ?) = ?
        WHERE h.hour <= ?
        GROUP BY h.hour
        ORDER BY h.hour ASC
        """
        
        # Extract world_id length for substring matching
        world_id_len = len(world_id)
        results = self.db.execute(query, (date_str, world_id_len, world_id, current_hour))
        return results

def print_location_history(db, date_str=None):
    """Print location history for a date."""
    query = VRCXQuery(db)
    locations = query.get_location_history(date_str)
    
    print(f"\n{'='*80}")
    print(f"Location History - {date_str or 'Today'}")
    print(f"{'='*80}")
    
    if not locations:
        print("No location data found for this date")
        return
    
    for loc in locations:
        hours = loc['duration_seconds'] / 3600 if loc['duration_seconds'] else 0
        print(f"[{loc['created_at']}] {loc['world_name']} ({loc['location']})")
        print(f"  Duration: {hours:.2f} hours ({loc['duration_seconds']}s)")
        print()


def print_hour_by_hour_summary(db, date_str=None, is_unique=False):
    """Print hour-by-hour summary of people in instances."""
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_by_hour(date_str)
        col = 'unique_visitors'
        title = 'Hour-by-Hour Summary (Unique Visitors)'
    else:
        summary = query.get_hour_by_hour_summary(date_str)
        col = 'unique_people'
        title = 'Hour-by-Hour Summary - All Visitors'
    
    print(f"\n{'='*50}")
    print(f"{title} - {date_str or 'Today'}")
    print(f"{'='*50}")
    
    if not summary:
        print("No data found for this date")
        return
    
    # Print header
    print(f"{'Hour':<6} {'People':<10}")
    print("-" * 50)
    
    for row in summary:
        hour = f"{row['hour']:02d}:00"
        people = row[col] or 0
        
        print(f"{hour:<6} {people:<10}")


def print_hour_by_hour_average(db, start_date_str=None, end_date_str=None, is_unique=False):
    """Print average hour-by-hour attendance across a date range."""
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_average(start_date_str, end_date_str)
        col = 'avg_unique_visitors'
        title = 'Average Attendance by Hour (Unique Visitors)'
    else:
        summary = query.get_hour_by_hour_average(start_date_str, end_date_str)
        col = 'avg_unique_people'
        title = 'Average Attendance by Hour - All Visitors'
    
    print(f"\n{'='*50}")
    print(f"{title} - {start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*50}")
    
    if not summary:
        print("No data found for this date range")
        return
    
    # Print header
    print(f"{'Hour':<6} {'Avg People':<15}")
    print("-" * 50)
    
    for row in summary:
        hour = f"{row['hour']:02d}:00"
        avg_people = row[col] or 0
        
        print(f"{hour:<6} {avg_people:<15}")


def print_daily_hourly_summary(db, start_date_str=None, end_date_str=None):
    """Print hour-by-hour summary for each day in a date range."""
    query = VRCXQuery(db)
    summary = query.get_daily_hourly_summary(start_date_str, end_date_str)
    
    print(f"\n{'='*60}")
    print(f"Daily Hour-by-Hour Attendance - {start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*60}")
    
    if not summary:
        print("No data found for this date range")
        return
    
    # Print header
    print(f"{'Date':<12} {'Hour':<6} {'People':<10}")
    print("-" * 60)
    
    current_date = None
    for row in summary:
        date = row['date']
        hour = f"{row['hour']:02d}:00"
        people = row['unique_people'] or 0
        
        # Add blank line between dates for readability
        if current_date and current_date != date:
            print()
        
        print(f"{date:<12} {hour:<6} {people:<10}")
        current_date = date


def print_day_of_week_average(db, start_date_str=None, end_date_str=None, is_unique=False):
    """Print average attendance by day of week."""
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_day_of_week(start_date_str, end_date_str)
        col = 'avg_unique_visitors'
        title = 'Average Attendance by Day of Week (Unique Visitors)'
    else:
        summary = query.get_day_of_week_average(start_date_str, end_date_str)
        col = 'avg_unique_people'
        title = 'Average Attendance by Day of Week - All Visitors'
    
    print(f"\n{'='*40}")
    print(f"{title}")
    print(f"{start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*40}")
    
    if not summary:
        print("No data found for this date range")
        return
    
    # Days of week (Sunday=0 to Saturday=6 in SQLite)
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    # Print header
    print(f"{'Day of Week':<12} {'Avg People':<15}")
    print("-" * 40)
    
    for row in summary:
        day_name = days[row['day_of_week']]
        avg_people = row[col] or 0
        
        print(f"{day_name:<12} {avg_people:<15}")


def print_world_list(db, start_date_str=None, end_date_str=None):
    """Print list of unique worlds visited during a date range."""
    query = VRCXQuery(db)
    worlds = query.get_unique_worlds(start_date_str, end_date_str)
    
    print(f"\n{'='*80}")
    print(f"Worlds Visited - {start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*80}")
    
    if not worlds:
        print("No world data found for this date range")
        return
    
    print(f"{'World Name':<40} {'World ID':<45} {'Days':<6} {'Events':<8}")
    print("-" * 99)
    
    for row in worlds:
        world_name = (row['world_name'] or 'Unknown')[:40]
        world_id = (row['world_id'] or 'Unknown')[:45]
        days = row['visit_count'] or 0
        events = row['total_events'] or 0
        
        print(f"{world_name:<40} {world_id:<45} {days:<6} {events:<8}")


def print_instances_for_world(db, world_id, world_name=None, start_date_str=None, end_date_str=None):
    """Print list of unique instances for a specific world."""
    query = VRCXQuery(db)
    instances = query.get_unique_instances_for_world(world_id, start_date_str, end_date_str)
    
    title = world_name or world_id
    print(f"\n{'='*130}")
    print(f"Instances for World: {title}")
    print(f"{start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*130}")
    
    if not instances:
        print("No instance data found for this world in the date range")
        return
    
    print(f"{'Instance Number':<100} {'Days':<6} {'Last Visited':<20}")
    print("-" * 130)
    
    for row in instances:
        instance_num = (row['instance_number'] or 'Unknown')[:100]
        days = row['visit_count'] or 0
        last_visited = row['last_visited'] or 'Unknown'
        
        print(f"{instance_num:<100} {days:<6} {last_visited:<20}")



def print_hour_by_hour_summary_for_world(db, world_id, world_name=None, date_str=None, is_unique=False):
    """Print hour-by-hour summary for a specific world on a given date."""
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_by_hour_for_world(world_id, date_str)
        col = 'unique_visitors'
        title = f'Hour-by-Hour Summary (Unique Visitors) - {world_name or world_id}'
    else:
        summary = query.get_hour_by_hour_summary_for_world(world_id, date_str)
        col = 'unique_people'
        title = f'Hour-by-Hour Summary - {world_name or world_id}'
    
    print(f"\n{'='*60}")
    print(f"{title} - {date_str or 'Today'}")
    print(f"{'='*60}")
    
    if not summary:
        print("No data found for this world on this date")
        return
    
    # Print header
    print(f"{'Hour':<6} {'People':<10}")
    print("-" * 60)
    
    for row in summary:
        hour = f"{row['hour']:02d}:00"
        people = row[col] or 0
        
        print(f"{hour:<6} {people:<10}")


def print_hour_by_hour_summary_for_instance(db, instance_id, date_str=None, is_unique=False):
    """Print hour-by-hour summary for a specific instance on a given date."""
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_by_hour_for_instance(instance_id, date_str)
        col = 'unique_visitors'
        title = f'Hour-by-Hour Summary (Unique Visitors)'
    else:
        summary = query.get_hour_by_hour_summary_for_instance(instance_id, date_str)
        col = 'unique_people'
        title = f'Hour-by-Hour Summary'
    
    if date_str is None:
        date_str = datetime.now().strftime('%Y-%m-%d')
    
    print(f"\n{'='*100}")
    print(f"{title} - {date_str}")
    print(f"Instance ID: {instance_id}")
    print(f"{'='*100}")
    
    if not summary:
        print("No data found for this instance on this date")
        return
    
    # Print header
    print(f"{'Hour':<6} {'People':<10}")
    print("-" * 100)
    
    for row in summary:
        hour = f"{row['hour']:02d}:00"
        people = row[col] or 0
        
        print(f"{hour:<6} {people:<10}")


def print_users_for_instance(db, instance_id, start_date_str=None, end_date_str=None):
    """Print list of unique users who visited a specific instance."""
    query = VRCXQuery(db)
    users = query.get_users_for_instance(instance_id, start_date_str, end_date_str)
    
    if start_date_str is None:
        start_date_str = datetime.now().strftime('%Y-%m-%d')
    if end_date_str is None:
        end_date_str = start_date_str
    
    print(f"\n{'='*130}")
    print(f"Users who visited instance: {instance_id}")
    print(f"Date range: {start_date_str} to {end_date_str}")
    print(f"{'='*130}")
    
    if not users:
        print("No users found for this instance in the date range")
        return
    
    # Calculate totals
    total_users = len(list(users))
    users = query.get_users_for_instance(instance_id, start_date_str, end_date_str)  # Re-query
    total_visits = sum(row['visit_count'] for row in users)
    users = query.get_users_for_instance(instance_id, start_date_str, end_date_str)  # Re-query again
    
    print(f"\nTotal Unique Users: {total_users}")
    print(f"Total Visits: {total_visits}\n")
    
    # Print header
    print(f"{'User Name':<30} {'Visits':<10} {'Days':<6} {'First Visit':<20} {'Last Visit':<20}")
    print("-" * 130)
    
    for row in users:
        user = (row['display_name'] or 'Unknown')[:30]
        # Encode safely to avoid Unicode errors on Windows terminals
        try:
            user_safe = user.encode('ascii', 'replace').decode('ascii')
        except:
            user_safe = 'Unknown'
        visits = row['visit_count'] or 0
        days = row['days_visited'] or 0
        first = row['first_visit'] or 'Unknown'
        last = row['last_visit'] or 'Unknown'
        
        try:
            print(f"{user_safe:<30} {visits:<10} {days:<6} {first:<20} {last:<20}")
        except UnicodeEncodeError:
            print(f"{'<encoding-error>':<30} {visits:<10} {days:<6} {first:<20} {last:<20}")


def print_day_of_week_average(db, start_date_str=None, end_date_str=None, is_unique=False):
    """Print attendance by day of week for each week in the date range."""
    query = VRCXQuery(db)
    summary = query.get_weekly_day_of_week_breakdown(start_date_str, end_date_str)
    
    if not summary:
        print("No data found for this date range")
        return
    
    print(f"\n{'='*60}")
    print(f"Weekly Breakdown by Day of Week")
    print(f"{start_date_str} to {end_date_str or start_date_str}")
    print(f"{'='*60}\n")
    
    # Group by week
    current_week = None
    for row in summary:
        week_key = (row['week_start'], row['week_end'])
        
        # Print week header when we encounter a new week
        if current_week != week_key:
            if current_week is not None:
                print()  # Blank line between weeks
            print(f"Week: {row['week_start']} to {row['week_end']}")
            print("-" * 40)
            current_week = week_key
        
        day_name = row['day_name']
        people = row['unique_people'] or 0
        print(f"  {day_name:<12} {people:>8}")
    
    print()


# ==============================================================================
# Chart Generation Functions
# ==============================================================================

def create_average_chart(db, output_file, start_date_str=None, end_date_str=None, chart_label='Average Hourly Attendance - All Visitors', is_unique=False):
    """Create a bar chart showing average attendance by hour."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return
    
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_average(start_date_str, end_date_str)
    else:
        summary = query.get_hour_by_hour_average(start_date_str, end_date_str)
    
    if not summary:
        print("No data to chart")
        return
    
    # Get the correct column name based on whether it's unique or not
    col = 'avg_unique_visitors' if is_unique else 'avg_unique_people'
    
    # Extract data and create a dict for easy lookup
    data_dict = {row['hour']: row[col] or 0 for row in summary}
    
    # Create arrays for all 24 hours, filling missing hours with 0
    all_hours = list(range(24))
    avg_people = [data_dict.get(h, 0) for h in all_hours]
    
    # Create chart
    fig, ax = plt.subplots(figsize=(10, 7))
    ax.bar(all_hours, avg_people, color='#1f77b4', width=0.8)
    ax.set_xlabel('Hour', fontsize=12)
    ax.set_ylabel('Average People', fontsize=12)
    ax.set_title('Average People by Hour', fontsize=14, fontweight='bold')
    ax.set_xticks(all_hours)
    ax.set_xticklabels([f'{h:02d}' for h in all_hours])
    ax.grid(axis='y', alpha=0.3)
    
    # Add info box at bottom
    fig.text(0.5, 0.02, chart_label, ha='center', fontsize=10, 
             bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
    
    plt.subplots_adjust(bottom=0.12)
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"[OK] Chart saved to {output_file}")


def create_daily_chart(db, output_file, date_str=None, chart_label='Hourly Attendance - All Visitors', is_unique=False):
    """Create a bar chart showing hourly attendance for a specific day."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return
    
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_by_hour(date_str)
        col = 'unique_visitors'
        y_label = 'Unique Visitors'
    else:
        summary = query.get_hour_by_hour_summary(date_str)
        col = 'unique_people'
        y_label = 'Unique People'
    
    if not summary:
        print("No data to chart")
        return
    
    # Extract data and create a dict for easy lookup
    data_dict = {row['hour']: row[col] or 0 for row in summary}
    
    # Create arrays for all 24 hours, filling missing hours with 0
    all_hours = list(range(24))
    people = [data_dict.get(h, 0) for h in all_hours]
    
    # Create chart
    fig, ax = plt.subplots(figsize=(10, 7))
    ax.bar(all_hours, people, color='#1f77b4', width=0.8)
    ax.set_xlabel('Hour', fontsize=12)
    ax.set_ylabel(y_label, fontsize=12)
    ax.set_title(f'Hourly Attendance - {date_str}', fontsize=14, fontweight='bold')
    ax.set_xticks(all_hours)
    ax.set_xticklabels([f'{h:02d}' for h in all_hours])
    ax.grid(axis='y', alpha=0.3)
    
    # Add info box at bottom
    fig.text(0.5, 0.02, chart_label, ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
    
    plt.subplots_adjust(bottom=0.12)
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"[OK] Chart saved to {output_file}")


def create_daily_chart_for_instance(db, output_file, instance_id, date_str=None, chart_label='Hourly Attendance - Instance', is_unique=False):
    """Create a bar chart showing hourly attendance for a specific instance."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return
    
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_by_hour_for_instance(instance_id, date_str)
        col = 'unique_visitors'
        y_label = 'Unique Visitors'
    else:
        summary = query.get_hour_by_hour_summary_for_instance(instance_id, date_str)
        col = 'unique_people'
        y_label = 'Unique People'
    
    if date_str is None:
        date_str = datetime.now().strftime('%Y-%m-%d')
    
    if not summary:
        print("No data to chart")
        return
    
    # Extract data and create a dict for easy lookup
    data_dict = {row['hour']: row[col] or 0 for row in summary}
    
    # Create arrays for all 24 hours, filling missing hours with 0
    all_hours = list(range(24))
    people = [data_dict.get(h, 0) for h in all_hours]
    
    # Create chart
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.bar(all_hours, people, color='#ff7f0e', width=0.8)
    ax.set_xlabel('Hour', fontsize=12)
    ax.set_ylabel(y_label, fontsize=12)
    ax.set_title(f'Instance Hourly Attendance - {date_str}\n{instance_id}', fontsize=12, fontweight='bold')
    ax.set_xticks(all_hours)
    ax.set_xticklabels([f'{h:02d}' for h in all_hours])
    ax.grid(axis='y', alpha=0.3)
    
    # Add info box at bottom
    fig.text(0.5, 0.02, chart_label, ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
    
    plt.subplots_adjust(bottom=0.15)
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"[OK] Chart saved to {output_file}")


def create_day_of_week_chart(db, output_file, start_date_str=None, end_date_str=None, chart_label='Day of Week - All Visitors', is_unique=False):
    """Create a bar chart showing average attendance by day of week."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return
    
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_day_of_week(start_date_str, end_date_str)
    else:
        summary = query.get_day_of_week_average(start_date_str, end_date_str)
    
    if not summary:
        print("No data to chart")
        return
    
    # Days of week (Sunday=0 to Saturday=6 in SQLite)
    days_full = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    # Extract data - handle both column names
    day_indices = [row['day_of_week'] for row in summary]
    day_names = [days_full[idx] for idx in day_indices]
    
    # Get the correct column name based on whether it's unique or not
    if is_unique:
        avg_people = [row['avg_unique_visitors'] or 0 for row in summary]
    else:
        avg_people = [row['avg_unique_people'] or 0 for row in summary]
    
    # Create chart
    fig, ax = plt.subplots(figsize=(10, 7))
    ax.bar(day_names, avg_people, color='#2ca02c', width=0.6)
    ax.set_xlabel('Day of Week', fontsize=12)
    ax.set_ylabel('Average Unique People', fontsize=12)
    ax.set_title(f'Average Attendance by Day of Week\n{start_date_str} to {end_date_str or start_date_str}', 
              fontsize=14, fontweight='bold')
    ax.tick_params(axis='x', rotation=45)
    ax.grid(axis='y', alpha=0.3)
    
    # Add info box at bottom
    fig.text(0.5, 0.02, chart_label, ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
    
    plt.subplots_adjust(bottom=0.15)
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"[OK] Chart saved to {output_file}")


def create_weekly_charts(db, output_dir, start_date_str=None, end_date_str=None, chart_label='Weekly Breakdown - All Visitors', is_unique=False):
    """Create separate bar charts for each week showing day-of-week attendance."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return []
    
    from pathlib import Path
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_weekly(start_date_str, end_date_str)
    else:
        summary = query.get_weekly_day_of_week_breakdown(start_date_str, end_date_str)
    
    if not summary:
        print("No data to chart")
        return []
    
    # Days of week for ordering
    days_full = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    # Group data by week
    weeks = {}
    for row in summary:
        week_key = (row['week_start'], row['week_end'])
        if week_key not in weeks:
            weeks[week_key] = {}
        # Get the correct column name (sqlite3.Row doesn't have .get() method)
        try:
            people_count = row['unique_visitors']
        except (KeyError, IndexError):
            try:
                people_count = row['unique_people']
            except (KeyError, IndexError):
                people_count = 0
        weeks[week_key][row['day_of_week']] = people_count or 0
    
    # Create a chart for each week
    chart_files = []
    for (week_start, week_end), week_data in sorted(weeks.items()):
        # Prepare data for this week
        day_indices = sorted(week_data.keys())
        day_names = [days_full[idx] for idx in day_indices]
        people = [week_data[idx] for idx in day_indices]
        
        # Create chart
        fig, ax = plt.subplots(figsize=(10, 7))
        ax.bar(day_names, people, color='#ff7f0e', width=0.6)
        ax.set_xlabel('Day of Week', fontsize=12)
        ax.set_ylabel('Unique People', fontsize=12)
        ax.set_title(f'Weekly Attendance: {week_start} to {week_end}', 
                  fontsize=14, fontweight='bold')
        ax.tick_params(axis='x', rotation=45)
        ax.grid(axis='y', alpha=0.3)
        
        # Add info box at bottom
        fig.text(0.5, 0.02, chart_label, ha='center', fontsize=10,
                 bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
        
        plt.subplots_adjust(bottom=0.15)
        
        # Save chart
        output_file = Path(output_dir) / f"vrcx_week_{week_start}_to_{week_end}.png"
        plt.savefig(output_file, dpi=150, bbox_inches='tight')
        plt.close()
        
        chart_files.append(str(output_file))
        print(f"[OK] Chart saved to {output_file}")
    
    return chart_files


def create_combined_weekly_chart(db, output_file, start_date_str=None, end_date_str=None, chart_label='Weekly Breakdown - All Visitors', is_unique=False):
    """Create a single combined chart with all weeks as subplots."""
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')  # Use non-interactive backend
    except ImportError:
        print("WARNING: matplotlib not installed. Install with: pip install matplotlib")
        return
    
    from pathlib import Path
    import math
    
    query = VRCXQuery(db)
    if is_unique:
        summary = query.get_unique_visitors_weekly(start_date_str, end_date_str)
    else:
        summary = query.get_weekly_day_of_week_breakdown(start_date_str, end_date_str)
    
    if not summary:
        print("No data to chart")
        return
    
    # Days of week for ordering
    days_full = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    # Group data by week
    weeks = {}
    for row in summary:
        week_key = (row['week_start'], row['week_end'])
        if week_key not in weeks:
            weeks[week_key] = {}
        # Get the correct column name (sqlite3.Row doesn't have .get() method)
        try:
            people_count = row['unique_visitors']
        except (KeyError, IndexError):
            try:
                people_count = row['unique_people']
            except (KeyError, IndexError):
                people_count = 0
        weeks[week_key][row['day_of_week']] = people_count or 0
    
    num_weeks = len(weeks)
    if num_weeks == 0:
        print("No weeks to chart")
        return
    
    # Calculate grid layout (prefer 2 columns, adjust rows as needed)
    cols = 2 if num_weeks > 1 else 1
    rows = math.ceil(num_weeks / cols)
    
    # Create figure with subplots (reserve space for info box at bottom)
    fig, axes = plt.subplots(rows, cols, figsize=(15, 5 * rows + 0.8))
    
    # Flatten axes array for easier iteration
    if num_weeks == 1:
        axes = [axes]
    else:
        axes = axes.flatten() if rows > 1 else [axes] if cols == 1 else axes
    
    # Plot each week in a subplot
    for idx, ((week_start, week_end), week_data) in enumerate(sorted(weeks.items())):
        ax = axes[idx]
        
        # Prepare data for this week
        day_indices = sorted(week_data.keys())
        day_names = [days_full[i] for i in day_indices]
        people = [week_data[i] for i in day_indices]
        
        # Create bar chart
        ax.bar(day_names, people, color='#ff7f0e', width=0.6)
        ax.set_title(f'{week_start} to {week_end}', fontsize=12, fontweight='bold')
        ax.set_xlabel('Day of Week', fontsize=10)
        ax.set_ylabel('Unique People', fontsize=10)
        ax.tick_params(axis='x', rotation=45)
        ax.grid(axis='y', alpha=0.3)
    
    # Hide any unused subplots
    for idx in range(num_weeks, len(axes)):
        axes[idx].set_visible(False)
    
    # Overall title
    fig.suptitle(f'Weekly Attendance Breakdown: {start_date_str} to {end_date_str or start_date_str}',
                 fontsize=16, fontweight='bold', y=0.99)
    
    # Add info box at bottom
    fig.text(0.5, 0.01, chart_label, ha='center', fontsize=10,
             bbox=dict(boxstyle='round', facecolor='#f0f0f0', edgecolor='#cccccc', pad=0.5))
    
    plt.subplots_adjust(bottom=0.08)
    plt.savefig(output_file, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"[OK] Combined chart saved to {output_file}")


def export_to_csv(db, output_file, date_str=None, start_date_str=None, end_date_str=None, is_average=False, is_daily=False, is_day_of_week=False, is_weekly=False, is_unique=False):
    """Export hour-by-hour data to CSV file."""
    try:
        import csv
    except ImportError:
        print("ERROR: csv module not available")
        return
    
    query = VRCXQuery(db)
    
    if is_unique:
        if is_weekly:
            summary = query.get_unique_visitors_weekly(start_date_str, end_date_str)
            fieldnames = ['Week Start', 'Week End', 'Day of Week', 'Unique Visitors']
        elif is_day_of_week:
            summary = query.get_unique_visitors_day_of_week(start_date_str, end_date_str)
            fieldnames = ['Day of Week', 'Avg Unique Visitors']
            days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        elif is_average:
            summary = query.get_unique_visitors_average(start_date_str, end_date_str)
            fieldnames = ['Hour', 'Avg Unique Visitors']
        elif is_daily:
            summary = query.get_unique_visitors_daily(start_date_str, end_date_str)
            fieldnames = ['Date', 'Unique Visitors']
        else:
            summary = query.get_unique_visitors_by_hour(date_str)
            fieldnames = ['Hour', 'Unique Visitors']
    else:
        if is_weekly:
            summary = query.get_weekly_day_of_week_breakdown(start_date_str, end_date_str)
            fieldnames = ['Week Start', 'Week End', 'Day of Week', 'People']
        elif is_day_of_week:
            summary = query.get_day_of_week_average(start_date_str, end_date_str)
            fieldnames = ['Day of Week', 'Avg People']
            days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        elif is_average:
            summary = query.get_hour_by_hour_average(start_date_str, end_date_str)
            fieldnames = ['Hour', 'Avg People']
        elif is_daily:
            summary = query.get_daily_hourly_summary(start_date_str, end_date_str)
            fieldnames = ['Date', 'Hour', 'People']
        else:
            summary = query.get_hour_by_hour_summary(date_str)
            fieldnames = ['Hour', 'People']
    
    if not summary:
        print("No data to export")
        return
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(fieldnames)
        
        for row in summary:
            if is_unique:
                if is_weekly:
                    writer.writerow([
                        row['week_start'],
                        row['week_end'],
                        row['day_name'],
                        row['unique_visitors'] or 0
                    ])
                elif is_day_of_week:
                    writer.writerow([
                        days[row['day_of_week']],
                        row['avg_unique_visitors'] or 0
                    ])
                elif is_average:
                    writer.writerow([
                        f"{row['hour']:02d}:00",
                        row['avg_unique_visitors'] or 0
                    ])
                elif is_daily:
                    writer.writerow([
                        row['date'],
                        row['unique_visitors'] or 0
                    ])
                else:
                    writer.writerow([
                        f"{row['hour']:02d}:00",
                        row['unique_visitors'] or 0
                    ])
            else:
                if is_weekly:
                    writer.writerow([
                        row['week_start'],
                        row['week_end'],
                        row['day_name'],
                        row['unique_people'] or 0
                    ])
                elif is_day_of_week:
                    writer.writerow([
                        days[row['day_of_week']],
                        row['avg_unique_people'] or 0
                    ])
                elif is_average:
                    writer.writerow([
                        f"{row['hour']:02d}:00",
                        row['avg_unique_people'] or 0
                    ])
                elif is_daily:
                    writer.writerow([
                        row['date'],
                        f"{row['hour']:02d}:00",
                        row['unique_people'] or 0
                    ])
                else:
                    writer.writerow([
                        f"{row['hour']:02d}:00",
                        row['unique_people'] or 0
                    ])
    
    print(f"[OK] Exported to {output_file}")


def export_to_csv_for_instance(db, output_file, instance_id, date_str=None, is_unique=False):
    """Export hour-by-hour data for a specific instance to CSV file."""
    try:
        import csv
    except ImportError:
        print("ERROR: csv module not available")
        return
    
    query = VRCXQuery(db)
    
    if date_str is None:
        date_str = datetime.now().strftime('%Y-%m-%d')
    
    if is_unique:
        summary = query.get_unique_visitors_by_hour_for_instance(instance_id, date_str)
        fieldnames = ['Hour', 'Unique Visitors']
    else:
        summary = query.get_hour_by_hour_summary_for_instance(instance_id, date_str)
        fieldnames = ['Hour', 'People']
    
    if not summary:
        print("No data to export")
        return
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        # Write instance ID as first row for context
        writer.writerow(['Instance ID', instance_id])
        writer.writerow(['Date', date_str])
        writer.writerow([])  # Blank row for separation
        writer.writerow(fieldnames)
        
        for row in summary:
            if is_unique:
                writer.writerow([
                    f"{row['hour']:02d}:00",
                    row['unique_visitors'] or 0
                ])
            else:
                writer.writerow([
                    f"{row['hour']:02d}:00",
                    row['unique_people'] or 0
                ])
    
    print(f"[OK] Exported to {output_file}")


def export_to_excel(db, output_file, date_str=None, start_date_str=None, end_date_str=None, is_average=False, is_daily=False, is_day_of_week=False, is_weekly=False, is_unique=False):
    """Export hour-by-hour data to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("WARNING: openpyxl not installed. Install with: pip install openpyxl")
        return
    
    query = VRCXQuery(db)
    
    if is_unique:
        if is_weekly:
            summary = query.get_unique_visitors_weekly(start_date_str, end_date_str)
            headers = ['Week Start', 'Week End', 'Day of Week', 'Unique Visitors']
        elif is_day_of_week:
            summary = query.get_unique_visitors_day_of_week(start_date_str, end_date_str)
            headers = ['Day of Week', 'Avg Unique Visitors']
            days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        elif is_average:
            summary = query.get_unique_visitors_average(start_date_str, end_date_str)
            headers = ['Hour', 'Avg Unique Visitors']
        elif is_daily:
            summary = query.get_unique_visitors_daily(start_date_str, end_date_str)
            headers = ['Date', 'Unique Visitors']
        else:
            summary = query.get_unique_visitors_by_hour(date_str)
            headers = ['Hour', 'Unique Visitors']
    else:
        if is_weekly:
            summary = query.get_weekly_day_of_week_breakdown(start_date_str, end_date_str)
            headers = ['Week Start', 'Week End', 'Day of Week', 'People']
        elif is_day_of_week:
            summary = query.get_day_of_week_average(start_date_str, end_date_str)
            headers = ['Day of Week', 'Avg People']
            days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        elif is_average:
            summary = query.get_hour_by_hour_average(start_date_str, end_date_str)
            headers = ['Hour', 'Avg People']
        elif is_daily:
            summary = query.get_daily_hourly_summary(start_date_str, end_date_str)
            headers = ['Date', 'Hour', 'People']
        else:
            summary = query.get_hour_by_hour_summary(date_str)
            headers = ['Hour', 'People']
    
    if not summary:
        print("No data to export")
        return
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hour-by-Hour"
    
    # Add headers
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row in summary:
        if is_unique:
            if is_weekly:
                ws.append([
                    row['week_start'],
                    row['week_end'],
                    row['day_name'],
                    row['unique_visitors'] or 0
                ])
            elif is_day_of_week:
                ws.append([
                    days[row['day_of_week']],
                    row['avg_unique_visitors'] or 0
                ])
            elif is_average:
                ws.append([
                    f"{row['hour']:02d}:00",
                    row['avg_unique_visitors'] or 0
                ])
            elif is_daily:
                ws.append([
                    row['date'],
                    row['unique_visitors'] or 0
                ])
            else:
                ws.append([
                    f"{row['hour']:02d}:00",
                    row['unique_visitors'] or 0
                ])
        else:
            if is_weekly:
                ws.append([
                    row['week_start'],
                    row['week_end'],
                    row['day_name'],
                    row['unique_people'] or 0
                ])
            elif is_day_of_week:
                ws.append([
                    days[row['day_of_week']],
                    row['avg_unique_people'] or 0
                ])
            elif is_average:
                ws.append([
                    f"{row['hour']:02d}:00",
                    row['avg_unique_people'] or 0
                ])
            elif is_daily:
                ws.append([
                    row['date'],
                    f"{row['hour']:02d}:00",
                    row['unique_people'] or 0
                ])
            else:
                ws.append([
                    f"{row['hour']:02d}:00",
                    row['unique_people'] or 0
                ])
    
    # Adjust column widths
    if is_weekly:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    elif is_day_of_week:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    elif is_average:
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    elif is_daily:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    else:
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        # Center align numeric columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
    
    wb.save(output_file)
    print(f"[OK] Exported to {output_file}")


# ==============================================================================
# Main
# ==============================================================================

def main():
        """Main entry point."""
        epilog = """
Modes (choose by flags):
    * Default / --date                 Hourly chart for a single date
    * --start-date --end-date          Hourly charts for each day in range
    * --average                        Average hourly across range
    * --day-of-week                    Average by day of week across range
    * --weekly                         Week-by-week day-of-week breakdown (also makes combined chart)
    * --instance-id / --instance       Hourly for a specific instance (single date or range)
    * --world-id                       Hourly for a specific world (single date or range)
    * --list-worlds                    List worlds visited in range
    * --list-instances                 List instances for a world in range (requires --world-id)

Common flags:
    * --unique         Count unique visitors (not total joins)
    * --export-data    Also write CSV/XLSX alongside charts
    * --verbose        Print table info
"""

    parser = argparse.ArgumentParser(
        description='Query VRCX database for VRChat activity analysis',
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=epilog)
    parser.add_argument('--date', type=str, help='Query a specific date (YYYY-MM-DD format)')
    parser.add_argument('--start-date', type=str, help='Start date for range query (YYYY-MM-DD format)')
    parser.add_argument('--end-date', type=str, help='End date for range query (YYYY-MM-DD format)')
    parser.add_argument('--average', action='store_true', help='Calculate average attendance across date range')
    parser.add_argument('--day-of-week', action='store_true', help='Show average attendance by day of week (Monday-Sunday)')
    parser.add_argument('--weekly', action='store_true', help='Show week-by-week breakdown with day-of-week attendance')
    parser.add_argument('--unique', action='store_true', help='Count unique visitors only once per day (ignores join/leave counts)')
    parser.add_argument('--export-data', action='store_true', help='Export data to CSV and Excel files')
    parser.add_argument('--verbose', action='store_true', help='Show verbose output including database table information')
    parser.add_argument('--list-worlds', action='store_true', help='List all worlds visited during the date range')
    parser.add_argument('--world-id', type=str, help='Filter reports to a specific world ID')
    parser.add_argument('--world-name', type=str, help='Optional display name for the world in reports')
    parser.add_argument('--list-instances', action='store_true', help='List all instances for a specific world (requires --world-id)')
    parser.add_argument('--instance-id', '--instance', type=str, dest='instance_id', help='Filter reports to a specific instance (full instance ID from location field)')
    
    args = parser.parse_args()
    
    # Determine dates to query
    is_date_range = args.start_date and args.end_date
    
    if args.weekly:
        if not args.start_date:
            args.start_date = datetime.now().strftime('%Y-%m-%d')
        if not args.end_date:
            args.end_date = args.start_date
    elif args.day_of_week:
        if not args.start_date:
            args.start_date = datetime.now().strftime('%Y-%m-%d')
        if not args.end_date:
            args.end_date = args.start_date
    elif args.average:
        if not args.start_date:
            args.start_date = datetime.now().strftime('%Y-%m-%d')
        if not args.end_date:
            args.end_date = args.start_date
    elif is_date_range:
        # Date range without average = daily breakdown
        pass
    else:
        if not args.date:
            args.date = datetime.now().strftime('%Y-%m-%d')
    
    # Connect to database
    db = VRCXDatabase(DATABASE_PATH)
    db.connect()
    
    try:
        # Show available tables (if verbose)
        if args.verbose:
            print(f"\nAvailable tables:")
            tables = db.get_table_names()
            for table in tables:
                print(f"  - {table}")
        
        # Run queries
        print("\n" + "="*80)
        print("QUERYING VRCX DATABASE")
        print("="*80)
        
        if args.list_worlds:
            # List all worlds in the date range
            if not args.start_date:
                args.start_date = args.date or datetime.now().strftime('%Y-%m-%d')
            if not args.end_date:
                args.end_date = args.start_date
            print_world_list(db, args.start_date, args.end_date)
        elif args.list_instances:
            # List all instances for a specific world
            if not args.world_id:
                print("Error: --list-instances requires --world-id")
                db.close()
                return
            if not args.start_date:
                args.start_date = args.date or datetime.now().strftime('%Y-%m-%d')
            if not args.end_date:
                args.end_date = args.start_date
            print_instances_for_world(db, args.world_id, args.world_name, args.start_date, args.end_date)
        elif args.instance_id:
            # Run reports filtered to a specific instance - show hourly attendance
            
            # If world_id is provided but instance_id doesn't contain it, prepend it
            if args.world_id and ':' not in args.instance_id:
                args.instance_id = f"{args.world_id}:{args.instance_id}"
            
            if not args.start_date:
                args.start_date = args.date or datetime.now().strftime('%Y-%m-%d')
            if not args.end_date:
                args.end_date = args.start_date
            
            # Generate hourly reports for instance (don't return early, let it continue to chart generation)
            if args.date:
                print_hour_by_hour_summary_for_instance(db, args.instance_id, args.date, args.unique)
            elif is_date_range:
                # For date ranges with instance filter, show hourly for each day
                print(f"\nHourly Attendance by Instance - {args.instance_id}")
                start = datetime.strptime(args.start_date, '%Y-%m-%d')
                end = datetime.strptime(args.end_date, '%Y-%m-%d')
                current = start
                
                while current <= end:
                    date_str = current.strftime('%Y-%m-%d')
                    print_hour_by_hour_summary_for_instance(db, args.instance_id, date_str, args.unique)
                    current += timedelta(days=1)
        elif args.world_id:
            # Run reports filtered to a specific world
            if args.world_name is None:
                args.world_name = args.world_id
            
            if args.date:
                print_hour_by_hour_summary_for_world(db, args.world_id, args.world_name, args.date, args.unique)
            elif is_date_range:
                # For date ranges with world filter, show hourly for each day
                print(f"\nHourly Attendance by World - {args.world_id}")
                start = datetime.strptime(args.start_date, '%Y-%m-%d')
                end = datetime.strptime(args.end_date, '%Y-%m-%d')
                current = start
                
                while current <= end:
                    date_str = current.strftime('%Y-%m-%d')
                    print_hour_by_hour_summary_for_world(db, args.world_id, args.world_name, date_str, args.unique)
                    current += timedelta(days=1)
        elif args.weekly:
            print_weekly_day_of_week_breakdown(db, args.start_date, args.end_date)
        elif args.day_of_week:
            print_day_of_week_average(db, args.start_date, args.end_date, args.unique)
        elif args.average:
            print_hour_by_hour_average(db, args.start_date, args.end_date, args.unique)
        elif is_date_range:
            print_daily_hourly_summary(db, args.start_date, args.end_date)
        else:
            print_location_history(db, args.date)
            print_hour_by_hour_summary(db, args.date, args.unique)
        
        # Skip chart generation if just listing worlds or instances
        if args.list_worlds or args.list_instances:
            print(f"\n[OK] List completed")
            db.close()
            return
        
        # Export charts (always generated)
        output_dir = Path(os.getenv('VRCX_REPORTS_OUTPUT_PATH', './vrcx_exports'))
        output_dir.mkdir(exist_ok=True)

        # Timestamp suffix so multiple runs on the same day don't overwrite
        run_ts = datetime.now().strftime('%Y%m%d-%H%M%S')

        print(f"\n{'='*80}")
        print("GENERATING CHARTS")
        print(f"{'='*80}")

        if args.weekly:
            filename_base = f"vrcx_weekly_{args.start_date}_to_{args.end_date}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            chart_label = "Weekly Breakdown - Unique Visitors" if args.unique else "Weekly Breakdown - All Visitors"
            chart_files = create_weekly_charts(db, str(output_dir), args.start_date, args.end_date, chart_label, args.unique)
            print(f"[OK] Created {len(chart_files)} individual weekly charts")
            combined_chart = output_dir / f"{filename_base}_combined.png"
            create_combined_weekly_chart(db, str(combined_chart), args.start_date, args.end_date, chart_label, args.unique)

            if args.export_data:
                csv_file = output_dir / f"{filename_base}.csv"
                xlsx_file = output_dir / f"{filename_base}.xlsx"
                export_to_csv(db, str(csv_file), start_date_str=args.start_date,
                             end_date_str=args.end_date, is_weekly=True, is_unique=args.unique)
                export_to_excel(db, str(xlsx_file), start_date_str=args.start_date,
                               end_date_str=args.end_date, is_weekly=True, is_unique=args.unique)

        elif args.day_of_week:
            filename_base = f"vrcx_day_of_week_{args.start_date}_to_{args.end_date}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            chart_label = "Day of Week - Unique Visitors" if args.unique else "Day of Week - All Visitors"
            chart_file = output_dir / f"{filename_base}.png"
            create_day_of_week_chart(db, str(chart_file), args.start_date, args.end_date, chart_label, args.unique)

            if args.export_data:
                csv_file = output_dir / f"{filename_base}.csv"
                xlsx_file = output_dir / f"{filename_base}.xlsx"
                export_to_csv(db, str(csv_file), start_date_str=args.start_date,
                             end_date_str=args.end_date, is_day_of_week=True, is_unique=args.unique)
                export_to_excel(db, str(xlsx_file), start_date_str=args.start_date,
                               end_date_str=args.end_date, is_day_of_week=True, is_unique=args.unique)

        elif args.average:
            filename_base = f"vrcx_average_{args.start_date}_to_{args.end_date}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            chart_label = "Average Hourly Attendance - Unique Visitors" if args.unique else "Average Hourly Attendance - All Visitors"
            chart_file = output_dir / f"{filename_base}.png"
            create_average_chart(db, str(chart_file), args.start_date, args.end_date, chart_label, args.unique)

            if args.export_data:
                csv_file = output_dir / f"{filename_base}.csv"
                xlsx_file = output_dir / f"{filename_base}.xlsx"
                export_to_csv(db, str(csv_file), start_date_str=args.start_date,
                             end_date_str=args.end_date, is_average=True, is_unique=args.unique)
                export_to_excel(db, str(xlsx_file), start_date_str=args.start_date,
                               end_date_str=args.end_date, is_average=True, is_unique=args.unique)

        elif args.instance_id:
            filename_base = f"vrcx_instance_{args.date or 'range'}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            chart_label = "Instance Hourly Attendance - Unique Visitors" if args.unique else "Instance Hourly Attendance - All Visitors"
            
            if args.date:
                # Single day for instance
                chart_file = output_dir / f"{filename_base}.png"
                create_daily_chart_for_instance(db, str(chart_file), args.instance_id, args.date, chart_label, args.unique)
                
                if args.export_data:
                    csv_file = output_dir / f"{filename_base}.csv"
                    export_to_csv_for_instance(db, str(csv_file), args.instance_id, args.date, args.unique)
            elif is_date_range:
                # Date range for instance - generate chart for each day
                start = datetime.strptime(args.start_date, '%Y-%m-%d')
                end = datetime.strptime(args.end_date, '%Y-%m-%d')
                current = start
                chart_count = 0
                
                while current <= end:
                    date_str = current.strftime('%Y-%m-%d')
                    chart_filename = f"vrcx_instance_{date_str}"
                    if args.unique:
                        chart_filename += "_unique"
                    chart_filename += f"_{run_ts}"
                    chart_file = output_dir / f"{chart_filename}.png"
                    create_daily_chart_for_instance(db, str(chart_file), args.instance_id, date_str, chart_label, args.unique)
                    chart_count += 1
                    current += timedelta(days=1)
                
                print(f"[OK] Created {chart_count} instance daily charts")
                
                if args.export_data:
                    csv_file = output_dir / f"vrcx_instance_{args.start_date}_to_{args.end_date}"
                    if args.unique:
                        csv_file += "_unique"
                    csv_file += f"_{run_ts}.csv"
                    # Export the most recent day for CSV when doing date range
                    export_to_csv_for_instance(db, str(csv_file), args.instance_id, args.end_date, args.unique)

        elif is_date_range:
            filename_base = f"vrcx_daily_{args.start_date}_to_{args.end_date}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            # Generate hourly charts for each day in the range
            chart_label = "Hourly Attendance - Unique Visitors" if args.unique else "Hourly Attendance - All Visitors"
            start = datetime.strptime(args.start_date, '%Y-%m-%d')
            end = datetime.strptime(args.end_date, '%Y-%m-%d')
            current = start
            chart_count = 0
            
            while current <= end:
                date_str = current.strftime('%Y-%m-%d')
                chart_filename = f"vrcx_hourly_{date_str}"
                if args.unique:
                    chart_filename += "_unique"
                chart_filename += f"_{run_ts}"
                chart_file = output_dir / f"{chart_filename}.png"
                create_daily_chart(db, str(chart_file), date_str, chart_label, args.unique)
                chart_count += 1
                current += timedelta(days=1)
            
            print(f"[OK] Created {chart_count} daily charts")

            if args.export_data:
                csv_file = output_dir / f"{filename_base}.csv"
                xlsx_file = output_dir / f"{filename_base}.xlsx"
                export_to_csv(db, str(csv_file), start_date_str=args.start_date,
                             end_date_str=args.end_date, is_daily=True, is_unique=args.unique)
                export_to_excel(db, str(xlsx_file), start_date_str=args.start_date,
                               end_date_str=args.end_date, is_daily=True, is_unique=args.unique)

        else:
            filename_base = f"vrcx_hourly_{args.date}"
            if args.unique:
                filename_base += "_unique"
            filename_base += f"_{run_ts}"

            chart_label = "Hourly Attendance - Unique Visitors" if args.unique else "Hourly Attendance - All Visitors"
            chart_file = output_dir / f"{filename_base}.png"
            create_daily_chart(db, str(chart_file), args.date, chart_label, args.unique)

            if args.export_data:
                csv_file = output_dir / f"{filename_base}.csv"
                xlsx_file = output_dir / f"{filename_base}.xlsx"
                export_to_csv(db, str(csv_file), date_str=args.date, is_unique=args.unique)
                export_to_excel(db, str(xlsx_file), date_str=args.date, is_unique=args.unique)

        # Completion message
        if args.export_data:
            print(f"\n[OK] All data exports completed in {output_dir}/")
        else:
            print(f"\n[OK] All charts completed in {output_dir}/")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        db.close()


if __name__ == '__main__':
    main()
